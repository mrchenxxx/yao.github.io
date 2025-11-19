import os
import time
import requests
from openpyxl import load_workbook

# 获取桌面路径
desktop = os.path.join(os.path.expanduser("~"), "Desktop")
file_path = os.path.join(desktop, "output.xlsx")

# 加载工作簿和活动工作表
wb = load_workbook(file_path)
ws = wb.active

# 遍历 B 列，从第2行开始
for row in range(2, ws.max_row + 1):
    cell = ws.cell(row=row, column=2)  # B列是第2列
    id_value = cell.value

    if not id_value:
        continue  # 跳过空单元格

    # 确保 id 是字符串（避免科学计数法等问题）
    if isinstance(id_value, (int, float)):
        # 尝试转为整数再转字符串，避免 Excel 中大数字变成浮点（如 1.92184962534239E+18）
        try:
            id_str = str(int(id_value))
        except (ValueError, OverflowError):
            id_str = str(id_value).strip()
    else:
        id_str = str(id_value).strip()

    url = f"https://zhongshiyao.vip/apis/event/api/yaodian/chapte/view?id={id_str}"
    print(f"正在请求 ID: {id_str} ...")

    try:
        response = requests.get(url, timeout=10)
        response.raise_for_status()
        json_data = response.json()

        if json_data.get("code") == 200 and "data" in json_data:
            data = json_data["data"]
            chapte_content = data.get("chapteContent", "")
            contrast_content = data.get("contrastContent", "")
            variation_analysis = data.get("variationAnalysis", "")

            # 写入 C、D、E 列（第3、4、5列）
            ws.cell(row=row, column=3, value=chapte_content)
            ws.cell(row=row, column=4, value=contrast_content)
            ws.cell(row=row, column=5, value=variation_analysis)
            print(f"✅ 成功写入 ID: {id_str}")
        else:
            print(f"⚠️ 请求无有效数据，ID: {id_str} | 响应: {json_data.get('msg', '未知错误')}")
    except Exception as e:
        print(f"❌ 请求失败 ID: {id_str}, 错误: {e}")

    # 防止被限流：每次请求后暂停 1 秒
    time.sleep(1.5)

# 保存回原文件
wb.save(file_path)
print("所有请求完成，文件已保存。")